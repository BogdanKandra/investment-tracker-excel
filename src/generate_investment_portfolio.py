import argparse
import json
from collections import defaultdict, deque
from calendar import monthrange
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string


# Import yfinance for real-time market data
try:
    import yfinance as yf
    YFINANCE_AVAILABLE = True
    print("✅ yfinance available - real-time data enabled")
except ImportError:
    YFINANCE_AVAILABLE = False
    print("⚠️  yfinance not available - using simulated data (run: pip install yfinance)")

# Import for currency conversion
try:
    import requests
    CURRENCY_API_AVAILABLE = True
    print("✅ requests available - currency conversion enabled")
except ImportError:
    CURRENCY_API_AVAILABLE = False
    print("⚠️  requests not available - using static exchange rates")

# Set paths
PROJECT_PATH = Path.cwd().parent
PORTFOLIOS_PATH = PROJECT_PATH / "results" / "portfolios"


def parse_arguments():
    """Parse command-line arguments"""
    parser = argparse.ArgumentParser(
        description='Generate investment portfolio Excel report from JSON data'
    )
    parser.add_argument(
        '--test',
        action='store_true',
        help='Use portfolio_test.json instead of portfolio.json'
    )
    return parser.parse_args()

def get_portfolio_json_path(use_test=False):
    """Get the portfolio JSON file path based on the test flag"""
    filename = "portfolio_test.json" if use_test else "portfolio.json"
    return PROJECT_PATH / "data" / filename

def get_currency_format(currency_symbol):
    """Helper function to get Excel currency format string"""
    if currency_symbol == '$':
        return '$#,##0.00'
    elif currency_symbol == '€':
        return '€#,##0.00'
    elif currency_symbol == '£':
        return '£#,##0.00'
    elif currency_symbol == '¥':
        return '¥#,##0'
    else:
        return f'"{currency_symbol}"#,##0.00'

def auto_adjust_column_widths(sheet):
    """Helper function to auto-adjust column widths while handling merged cells"""
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = None
        
        for cell in column_cells:
            # Skip merged cells that don't have column_letter attribute
            if hasattr(cell, 'column_letter'):
                column_letter = cell.column_letter
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        
        if column_letter:
            adjusted_width = min(max_length + 2, 25)
            sheet.column_dimensions[column_letter].width = adjusted_width

def get_stock_info_yfinance(symbol):
    """Get comprehensive stock information using yfinance"""
    if not YFINANCE_AVAILABLE:
        return None
    
    try:
        stock = yf.Ticker(symbol)
        info = stock.info
        hist = stock.history(period="1y")  # Get 1 year of data for 52-week calculations
        
        if hist.empty:
            print(f"No data available for {symbol}")
            return None
        
        return {
            'symbol': symbol,
            'company_name': info.get('longName', info.get('shortName', symbol)),
            'country': info.get('country', 'N/A'),
            'current_price': hist['Close'].iloc[-1],
            'previous_close': info.get('previousClose', 'N/A'),
            'open': info.get('open', 'N/A'),
            'day_high': info.get('dayHigh', 'N/A'),
            'day_low': info.get('dayLow', 'N/A'),
            '52_week_high': hist['High'].max(),
            '52_week_low': hist['Low'].min(),
            'volume': info.get('volume', 'N/A'),
            'market_cap': info.get('marketCap', 'N/A'),
            'pe_ratio': info.get('trailingPE', 'N/A'),
            'dividend_yield': info.get('dividendYield', 'N/A'),
            'sector': info.get('sector', 'N/A'),
            'industry': info.get('industry', 'N/A'),
            'currency': info.get('currency', 'N/A'),
            'quote_type': info.get('quoteType', 'N/A'),
            'last_updated': datetime.now().isoformat()
        }
    except Exception as e:
        print(f"Error fetching data for {symbol}: {e}")
        return None

def get_multiple_stocks_info(symbols):
    """Get information for multiple stocks efficiently"""
    if not YFINANCE_AVAILABLE:
        return {}
    
    print(f"Fetching real-time data for {len(symbols)} symbols...")
    stocks_info = {}
    
    for i, symbol in enumerate(symbols, 1):
        print(f"  [{i}/{len(symbols)}] Fetching {symbol}...")
        info = get_stock_info_yfinance(symbol)
        if info:
            stocks_info[symbol] = info
        else:
            print(f"  ⚠️  Failed to get data for {symbol}")
    
    print(f"✅ Successfully fetched data for {len(stocks_info)}/{len(symbols)} symbols")
    return stocks_info

def get_asset_class(quote_type, country):
    """Determine asset class based on quote_type and country"""
    if not quote_type or quote_type == 'N/A':
        return 'Unknown'
    
    if quote_type == "EQUITY":
        if country == "United States":
            return "US Stocks"
        elif country == "Romania":
            return "Romanian Stocks"
        else:
            return "International Stocks"
    elif quote_type == "ETF":
        return "ETF"
    elif quote_type == "CRYPTOCURRENCY":
        return "Crypto"
    else:
        return "Unknown"

def get_exchange_rates():
    """Get current exchange rates to USD and EUR"""
    exchange_rates = {
        'USD': 1.0,  # Base currency
        'EUR': 1.10,  # Default fallback rate (to USD)
        'RON': 0.22,  # Default fallback rate (to USD)
        '£': 1.25,   # Default fallback rate (to USD)
        '¥': 0.007,  # Default fallback rate (to USD)
        'EUR_RATES': {  # EUR-based rates (to EUR)
            'USD': 0.91,  # Default fallback
            'EUR': 1.0,   # Base
            'RON': 0.20,  # Default fallback
            'GBP': 1.14,  # Default fallback
            'JPY': 0.0064 # Default fallback
        }
    }
    
    if not CURRENCY_API_AVAILABLE:
        print("⚠️  Using static exchange rates (requests not available)")
        return exchange_rates
    
    try:
        # Fetch USD-based rates
        response_usd = requests.get("https://api.exchangerate-api.com/v4/latest/USD", timeout=5)
        if response_usd.status_code == 200:
            data_usd = response_usd.json()
            rates_usd = data_usd.get('rates', {})
            
            # Convert from USD-base to USD-target (invert rates)
            if 'EUR' in rates_usd:
                exchange_rates['EUR'] = 1.0 / rates_usd['EUR']
            if 'RON' in rates_usd:
                exchange_rates['RON'] = 1.0 / rates_usd['RON']
            if 'GBP' in rates_usd:
                exchange_rates['£'] = 1.0 / rates_usd['GBP']
            if 'JPY' in rates_usd:
                exchange_rates['¥'] = 1.0 / rates_usd['JPY']
        
        # Fetch EUR-based rates
        response_eur = requests.get("https://api.exchangerate-api.com/v4/latest/EUR", timeout=5)
        if response_eur.status_code == 200:
            data_eur = response_eur.json()
            rates_eur = data_eur.get('rates', {})
            
            # Direct rates from EUR (no inversion needed)
            eur_rates = {}
            if 'USD' in rates_eur:
                eur_rates['USD'] = 1.0 / rates_eur['USD']  # Invert to get EUR rate
            if 'RON' in rates_eur:
                eur_rates['RON'] = 1.0 / rates_eur['RON']  # Invert to get EUR rate
            if 'GBP' in rates_eur:
                eur_rates['GBP'] = 1.0 / rates_eur['GBP']  # Invert to get EUR rate
            if 'JPY' in rates_eur:
                eur_rates['JPY'] = 1.0 / rates_eur['JPY']   # Invert to get EUR rate
            
            eur_rates['EUR'] = 1.0  # EUR to EUR
            exchange_rates['EUR_RATES'] = eur_rates
        
        print(f"✅ Live exchange rates fetched:")
        print(f"   USD rates: EUR={exchange_rates['EUR']:.4f}, RON={exchange_rates['RON']:.4f}")
        print(f"   EUR rates: USD={exchange_rates['EUR_RATES']['USD']:.4f}, RON={exchange_rates['EUR_RATES']['RON']:.4f}")
                
    except Exception as e:
        print(f"⚠️  Exchange rate fetch failed: {e}, using static rates")
    
    return exchange_rates

def convert_to_usd(value, from_currency, exchange_rates):
    """Convert a value from any currency to USD"""
    if not value or value == 0:
        return 0
    
    # Handle currency symbol mapping
    currency_map = {
        '$': 'USD',
        '€': 'EUR', 
        'RON': 'RON',
        '£': 'GBP',
        '¥': 'JPY'
    }
    
    currency_code = currency_map.get(from_currency, from_currency)
    
    # Handle direct currency codes
    if currency_code == 'USD':
        return value  # Already in USD
    
    exchange_rate = exchange_rates.get(currency_code, 1.0)
    
    return value * exchange_rate

def convert_to_eur(value, from_currency, exchange_rates):
    """Convert a value from any currency to EUR"""
    if not value or value == 0:
        return 0
    
    # Handle currency symbol mapping
    currency_map = {
        '$': 'USD',
        '€': 'EUR', 
        'RON': 'RON',
        '£': 'GBP',
        '¥': 'JPY'
    }
    
    currency_code = currency_map.get(from_currency, from_currency)
    
    # Handle direct currency codes
    if currency_code == 'EUR':
        return value  # Already in EUR
    
    # Get EUR-based exchange rates (stored as EUR_RATES in exchange_rates)
    eur_rates = exchange_rates.get('EUR_RATES', {})
    exchange_rate = eur_rates.get(currency_code, 1.0)
    
    return value * exchange_rate

def create_global_holdings_with_usd_conversion(global_holdings_data, exchange_rates):
    """Create USD-converted version of global holdings for the global portfolio section"""
    usd_holdings = {}
    
    for symbol, data in global_holdings_data.items():
        original_currency = data.get('currency', '$')
        
        # Convert all financial values to USD
        usd_data = data.copy()
        usd_data['currency'] = '$'  # Set to USD
        usd_data['original_currency'] = original_currency  # Keep track of original
        
        # Convert weighted average cost
        usd_data['weighted_avg_cost'] = convert_to_usd(
            data['weighted_avg_cost'], original_currency, exchange_rates
        )
        
        # Convert current price if available
        if 'current_price' in data:
            usd_data['current_price'] = convert_to_usd(
                data['current_price'], original_currency, exchange_rates
            )
        
        # Recalculate total cost in USD
        usd_data['total_cost'] = usd_data['total_shares'] * usd_data['weighted_avg_cost']
        
        usd_holdings[symbol] = usd_data
        
    return usd_holdings

def create_global_holdings_with_eur_conversion(global_holdings_data, exchange_rates):
    """Create EUR-converted version of global holdings for the global portfolio section"""
    eur_holdings = {}
    
    for symbol, data in global_holdings_data.items():
        original_currency = data.get('currency', '$')
        
        # Convert all financial values to EUR
        eur_data = data.copy()
        eur_data['currency'] = '€'  # Set to EUR
        eur_data['original_currency'] = original_currency  # Keep track of original
        
        # Convert weighted average cost
        eur_data['weighted_avg_cost'] = convert_to_eur(
            data['weighted_avg_cost'], original_currency, exchange_rates
        )
        
        # Convert current price if available
        if 'current_price' in data:
            eur_data['current_price'] = convert_to_eur(
                data['current_price'], original_currency, exchange_rates
            )
        
        # Recalculate total cost in EUR
        eur_data['total_cost'] = eur_data['total_shares'] * eur_data['weighted_avg_cost']
        
        eur_holdings[symbol] = eur_data
        
    return eur_holdings

def load_portfolio_data(filename):
    """Load and process portfolio data from JSON file"""
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except FileNotFoundError:
        raise FileNotFoundError(f"Portfolio data file not found at: {filename}")
    except Exception as e:
        raise Exception(f"Error loading portfolio data from {filename}: {e}")
        
    # Process transactions per account and globally
    all_transactions = []
    accounts_data = {}
    
    for account in data.get('accounts', []):
        account_name = account.get('account_name')
        accounts_data[account_name] = {
            'transactions': account.get('transactions', []),
            'cash': account.get('cash', 0),  # Add cash field
            'currency': account.get('currency', '$'),  # Add currency field
            'holdings': defaultdict(lambda: {
                'name': '',
                'lots': deque(),  # FIFO queue of purchase lots: [(shares, price, date), ...]
                'total_shares': 0,
                'weighted_avg_cost': 0,
                'total_cost': 0,
                'sector': 'Unknown',
                'currency': '$'  # Default currency
            })
        }
        
        # Add to global transactions list
        for transaction in account.get('transactions', []):
            all_transactions.append(transaction)
    
    # Calculate current holdings by aggregating transactions using FIFO (global)
    global_holdings_data = defaultdict(lambda: {
        'name': '',
        'lots': deque(),  # FIFO queue of purchase lots: [(shares, price, date), ...]
        'total_shares': 0,
        'weighted_avg_cost': 0,
        'total_cost': 0,
        'sector': 'Unknown',
        'currency': '$'  # Default currency
    })
    
    def calculate_totals(symbol_data):
        """Helper function to recalculate totals from remaining lots"""
        total_shares = 0
        total_cost = 0
        
        for shares, price, _ in symbol_data['lots']:
            total_shares += shares
            total_cost += shares * price
        
        symbol_data['total_shares'] = total_shares
        symbol_data['total_cost'] = total_cost
        symbol_data['weighted_avg_cost'] = total_cost / total_shares if total_shares > 0 else 0
    
    def process_transactions(holdings_dict, transactions_list):
        """Process transactions for a given holdings dictionary"""
        for transaction in transactions_list:
            symbol = transaction['symbol']
            shares = transaction['shares']
            price = transaction['price']
            trans_type = transaction['type'].lower()
            date = transaction['date']
            currency = transaction.get('currency', '$')  # Default to $ if not specified

            if trans_type == 'buy':
                # Add new lot to the end of the queue (FIFO)
                holdings_dict[symbol]['lots'].append((shares, price, date))
                holdings_dict[symbol]['name'] = transaction['name']
                holdings_dict[symbol]['currency'] = currency  # Track currency
                
            elif trans_type == 'sell':
                # Use FIFO to sell shares from oldest lots first
                remaining_to_sell = shares
                
                while remaining_to_sell > 0 and holdings_dict[symbol]['lots']:
                    # Get the oldest lot (front of queue)
                    lot_shares, lot_price, lot_date = holdings_dict[symbol]['lots'][0]
                    
                    if lot_shares <= remaining_to_sell:
                        # Sell entire lot
                        remaining_to_sell -= lot_shares
                        holdings_dict[symbol]['lots'].popleft()  # Remove entire lot
                    else:
                        # Partially sell from this lot
                        new_lot_shares = lot_shares - remaining_to_sell
                        holdings_dict[symbol]['lots'][0] = (new_lot_shares, lot_price, lot_date)
                        remaining_to_sell = 0

            # Recalculate totals
            calculate_totals(holdings_dict[symbol])

        # Remove holdings with zero or negative shares
        return {k: v for k, v in holdings_dict.items() if v['total_shares'] > 0}
    
    # Process transactions for each account
    for account_name, account_data in accounts_data.items():
        account_data['holdings'] = process_transactions(account_data['holdings'], account_data['transactions'])
    
    # Sort all transactions by date before processing globally (crucial for FIFO)
    try:
        all_transactions.sort(key=lambda x: datetime.strptime(x['date'], "%d-%m-%Y"))
    except:
        # If date parsing fails, keep original order
        pass
    
    # Process all transactions globally
    global_holdings_data = process_transactions(global_holdings_data, all_transactions)
    
    # Get watchlist data
    watchlist_data = data.get('watchlist', [])

    # Get updated_at date
    updated_at = data.get('updated_at', datetime.now().strftime("%m-%d-%Y"))
    
    # NEW: Fetch real-time market data for all holdings
    all_symbols = set(global_holdings_data.keys())
    for account_data in accounts_data.values():
        all_symbols.update(account_data['holdings'].keys())
    
    # Also include watchlist symbols
    for item in watchlist_data:
        all_symbols.add(item['symbol'])
    
    real_time_data = {}
    if all_symbols and YFINANCE_AVAILABLE:
        print(f"\n🔄 Fetching real-time market data...")
        real_time_data = get_multiple_stocks_info(list(all_symbols))
        
        # Enrich global holdings with real-time data
        for symbol, holding in global_holdings_data.items():
            if symbol in real_time_data:
                rt_data = real_time_data[symbol]
                holding['current_price'] = rt_data['current_price']
                holding['sector'] = rt_data['sector']
                holding['company_name'] = rt_data['company_name']
                holding['market_cap'] = rt_data['market_cap']
                holding['pe_ratio'] = rt_data['pe_ratio']
                holding['dividend_yield'] = rt_data['dividend_yield']
                holding['52_week_high'] = rt_data['52_week_high']
                holding['52_week_low'] = rt_data['52_week_low']
                holding['asset_class'] = get_asset_class(rt_data['quote_type'], rt_data['country'])
                holding['data_source'] = 'yfinance'
        
        # Enrich account holdings with real-time data
        for account_name, account_data in accounts_data.items():
            for symbol, holding in account_data['holdings'].items():
                if symbol in real_time_data:
                    rt_data = real_time_data[symbol]
                    holding['current_price'] = rt_data['current_price']
                    holding['sector'] = rt_data['sector']
                    holding['company_name'] = rt_data['company_name']
                    holding['market_cap'] = rt_data['market_cap']
                    holding['pe_ratio'] = rt_data['pe_ratio']
                    holding['dividend_yield'] = rt_data['dividend_yield']
                    holding['52_week_high'] = rt_data['52_week_high']
                    holding['52_week_low'] = rt_data['52_week_low']
                    holding['asset_class'] = get_asset_class(rt_data['quote_type'], rt_data['country'])
                    holding['data_source'] = 'yfinance'

    # Fetch current exchange rates for global portfolio USD conversion
    print("\n💱 Fetching currency exchange rates...")
    exchange_rates = get_exchange_rates()
    
    # Get target asset class distribution
    target_asset_class_distribution = data.get('target_asset_class_distribution', {})
    
    return {
        'transactions': all_transactions,
        'holdings': global_holdings_data,
        'accounts': accounts_data,
        'watchlist': watchlist_data,
        'real_time_data': real_time_data,
        'exchange_rates': exchange_rates,
        'updated_at': updated_at,
        'target_asset_class_distribution': target_asset_class_distribution
    }

def create_investment_portfolio_template(portfolio_json_path):
    """
    Creates a comprehensive Excel template for personal investment portfolio management
    with multiple dynamic sheets containing realistic investment data.
    
    Args:
        portfolio_json_path: Path to the portfolio JSON file
    """
    try:
        # Load portfolio data from JSON
        print(f"\n📂 Loading portfolio data from: {portfolio_json_path}")
        portfolio_data = load_portfolio_data(portfolio_json_path)
        print(f"✅ Portfolio data loaded successfully")
        
        # Create a new workbook
        wb = Workbook()
        
        # Remove the default sheet
        wb.remove(wb.active)
        
        # Define styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")  # Dark green for investments
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        print("\n📊 Creating Excel sheets...")
        
        # 1. Overview Sheet
        overview_sheet = wb.create_sheet("Overview")
        create_portfolio_overview(overview_sheet, header_font, header_fill, border, portfolio_data)
        print("  ✓ Overview sheet created")
        
        # 2. Holdings Sheet
        holdings_sheet = wb.create_sheet("Holdings")
        create_holdings(holdings_sheet, header_font, header_fill, border, portfolio_data)
        print("  ✓ Holdings sheet created")
        
        # 3. Transactions Sheet
        transactions_sheet = wb.create_sheet("Transactions")
        create_transactions_history(transactions_sheet, header_font, header_fill, border, portfolio_data)
        print("  ✓ Transactions sheet created")
        
        # 4. Dividends Sheet
        dividend_sheet = wb.create_sheet("Dividends")
        create_dividend_tracker(dividend_sheet, header_font, header_fill, border, portfolio_data)
        print("  ✓ Dividends sheet created")
        
        # 5. Analysis Sheet
        performance_sheet = wb.create_sheet("Analysis")
        create_performance_analysis(performance_sheet, header_font, header_fill, border, portfolio_data)
        print("  ✓ Analysis sheet created")
        
        # 6. Watchlist Sheet
        watchlist_sheet = wb.create_sheet("Watchlist")
        create_watchlist(watchlist_sheet, header_font, header_fill, border, portfolio_data)
        print("  ✓ Watchlist sheet created")
        
        # Save the workbook
        print("\n💾 Saving Excel file...")
        # take the updated_at date and format it to %m-%d-%Y
        updated_at = portfolio_data['updated_at']
        updated_at = datetime.strptime(updated_at, "%d-%m-%Y").strftime("%Y-%m-%d")
        workbook_name = f"Investment_Portfolio_{updated_at}.xlsx"
        wb.save(PORTFOLIOS_PATH / workbook_name)
        print(f"✅ Investment portfolio '{workbook_name}' has been created successfully!")
        print(f"📁 Location: {PORTFOLIOS_PATH / workbook_name}")
    except FileNotFoundError as e:
        print(f"\n❌ ERROR: {e}")
        print(f"   Please ensure the portfolio data file exists at: {portfolio_json_path}")
        raise
    except Exception as e:
        print(f"\n❌ ERROR: An unexpected error occurred while creating the portfolio:")
        print(f"   {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        raise

def create_holdings(sheet, header_font, header_fill, border, portfolio_data):
    """Create holdings information with Excel formulas - multiple sections per account"""
    
    def create_holdings_section(start_row, section_title, holdings_dict, is_global=False, section_summary_rows=None, cash_amount=0, cash_currency='$'):
        """Create a holdings section starting at the specified row"""
        current_row = start_row
        data_rows = []  # Track actual data rows for conditional formatting
        
        # Section title - Extended to cover Asset Class column
        sheet.merge_cells(f'A{current_row}:K{current_row}')
        title_cell = sheet[f'A{current_row}']
        
        # Add data source indicator if yfinance is available
        data_source_indicator = ""
        if YFINANCE_AVAILABLE:
            data_source_indicator = " 📊 LIVE DATA"
        
        title_cell.value = section_title + data_source_indicator
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1
        
        # Headers - Added Asset Class column for Overview integration
        headers = ["Symbol", "Company Name", "Shares", "Avg Cost", "Current Price", 
                   "Market Value", "Unrealized Gain/Loss", "% Gain/Loss", "% of Portfolio", "Sector", "Asset Class"]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        current_row += 1
        
        # Holdings data
        holdings_rows = []
        for symbol, data in holdings_dict.items():
            # Use real-time price if available, otherwise use weighted average cost
            if 'current_price' in data and data.get('data_source') == 'yfinance':
                current_price = data['current_price']
                quote_type = data.get('quote_type', 'Unknown')
                country = data.get('country', 'Unknown')
                sector = data.get('sector', 'Unknown')
                company_name = data.get('company_name', data['name'])
            else:
                # Use weighted average cost if real-time price not available
                current_price = data['weighted_avg_cost']
                quote_type = "EQUITY"
                country = "United States"
                sector = data.get('sector', 'Unknown')
                company_name = data['name']
            
            currency = data.get('currency', '$')
            
            # Get asset class for Overview integration  
            asset_class = data.get('asset_class', get_asset_class(quote_type, country))
            
            # Update sector based on asset class
            if asset_class == "ETF":
                sector = "ETF"
            elif asset_class == "Crypto":
                sector = "Crypto"
            
            holdings_rows.append((symbol, company_name, data['total_shares'], 
                                data['weighted_avg_cost'], current_price, sector, currency, asset_class))
        
        # Calculate summary row number (where total portfolio value will be)
        # summary_row = start_row + 1 (title) + 1 (headers) + len(holdings_rows) + (1 if cash)
        summary_row = start_row + 2 + len(holdings_rows) + (1 if cash_amount > 0 else 0)
        
        # Add holdings data to sheet
        for symbol, company, shares, avg_cost, current_price, sector, currency, asset_class in holdings_rows:
            # Get currency formatting
            currency_format = get_currency_format(currency)
            market_value_format = currency_format.replace('.00', '')  # Remove decimals for market value
            
            # Input the basic data
            sheet.cell(row=current_row, column=1, value=symbol).border = border
            sheet.cell(row=current_row, column=2, value=company).border = border
            sheet.cell(row=current_row, column=3, value=shares).border = border
            sheet.cell(row=current_row, column=4, value=avg_cost).border = border
            sheet.cell(row=current_row, column=5, value=current_price).border = border
            sheet.cell(row=current_row, column=10, value=sector).border = border
            sheet.cell(row=current_row, column=11, value=asset_class).border = border
            
            # Market Value = Shares * Current Price
            market_value_cell = sheet.cell(row=current_row, column=6, value=f"=C{current_row}*E{current_row}")
            market_value_cell.border = border
            market_value_cell.number_format = market_value_format
            
            # Unrealized Gain/Loss = Market Value - (Shares * Avg Cost)
            gain_loss_cell = sheet.cell(row=current_row, column=7, value=f"=F{current_row}-(C{current_row}*D{current_row})")
            gain_loss_cell.border = border
            gain_loss_cell.number_format = currency_format
            
            # % Gain/Loss = Unrealized Gain/Loss / (Shares * Avg Cost)
            pct_gain_cell = sheet.cell(row=current_row, column=8, value=f"=G{current_row}/(C{current_row}*D{current_row})")
            pct_gain_cell.border = border
            pct_gain_cell.number_format = '0.00%'
            
            # % of Portfolio = Market Value / Total Portfolio Value (from summary row)
            pct_portfolio_cell = sheet.cell(row=current_row, column=9, 
                                            value=f"=F{current_row}/F${summary_row}")

            pct_portfolio_cell.border = border
            pct_portfolio_cell.number_format = '0.00%'
            
            # Additional formatting with currency
            sheet.cell(row=current_row, column=1).font = Font(bold=True)
            sheet.cell(row=current_row, column=4).number_format = currency_format  # Avg Cost
            sheet.cell(row=current_row, column=5).number_format = currency_format  # Current Price
            
            # Track this row for conditional formatting
            data_rows.append(current_row)
            
            current_row += 1
        
        # Add cash position row if cash > 0
        if cash_amount > 0:
            # Get currency formatting for cash
            cash_currency_format = get_currency_format(cash_currency)
            
            # Add cash row
            sheet.cell(row=current_row, column=1, value="CASH").border = border
            sheet.cell(row=current_row, column=1).font = Font(bold=True, color="008000")  # Green for cash
            sheet.cell(row=current_row, column=2, value="Cash Position").border = border
            sheet.cell(row=current_row, column=3, value="").border = border  # No shares for cash
            sheet.cell(row=current_row, column=4, value="").border = border  # No avg cost for cash
            sheet.cell(row=current_row, column=5, value="").border = border  # No current price for cash
            
            # Market Value = Cash Amount
            cash_market_value_cell = sheet.cell(row=current_row, column=6, value=cash_amount)
            cash_market_value_cell.border = border
            cash_market_value_cell.number_format = cash_currency_format.replace('.00', '')  # Remove decimals for cash
            cash_market_value_cell.font = Font(color="008000")  # Green for cash
            
            # No unrealized gain/loss for cash
            sheet.cell(row=current_row, column=7, value=0).border = border
            sheet.cell(row=current_row, column=8, value=0).border = border
            sheet.cell(row=current_row, column=8).number_format = '0.00%'
            
            # % of Portfolio calculation - will be computed in summary row calculation below
            cash_pct_row = current_row
            pct_portfolio_cell = sheet.cell(row=current_row, column=9, value="")  # Will be set after summary calculation
            pct_portfolio_cell.border = border
            pct_portfolio_cell.number_format = '0.00%'
            
            # Sector and Asset Class for cash
            sheet.cell(row=current_row, column=10, value="Cash").border = border
            sheet.cell(row=current_row, column=11, value="Cash").border = border
            
            # Track this row for conditional formatting (though cash won't have gains/losses)
            data_rows.append(current_row)
            
            current_row += 1
        else:
            cash_pct_row = None
        
        # Add summary row for the section
        if holdings_rows or cash_amount > 0:
            summary_row = current_row
            
            # Get currency format from first holding, cash, or use default
            if holdings_rows:
                summary_currency = holdings_rows[0][6]  # currency from first row
            else:
                summary_currency = cash_currency  # Use cash currency if no holdings
            summary_currency_format = get_currency_format(summary_currency)
            
            # Summary labels
            sheet.cell(row=summary_row, column=1, value="TOTALS / AVERAGES").font = Font(bold=True)
            sheet.cell(row=summary_row, column=1).border = border
            
            # Market Value Total (Column F)
            data_start = start_row + 2  # After title and headers
            data_end = summary_row - 1   # Before summary row
            market_value_total = sheet.cell(row=summary_row, column=6, 
                                          value=f"=SUM(F{data_start}:F{data_end})")
            market_value_total.border = border
            market_value_total.number_format = summary_currency_format.replace('.00', '')  # Remove decimals for totals
            market_value_total.font = Font(bold=True)
            
            # Unrealized Gain/Loss Total (Column G)
            unrealized_total = sheet.cell(row=summary_row, column=7, 
                                        value=f"=SUM(G{data_start}:G{data_end})")
            unrealized_total.border = border
            unrealized_total.number_format = summary_currency_format
            unrealized_total.font = Font(bold=True)
            
            # Average % Gain/Loss (Column H) = Total Unrealized Gain/Loss / Total Market Value
            avg_gain_loss = sheet.cell(row=summary_row, column=8, 
                                     value=f"=IF(F{summary_row}<>0,G{summary_row}/F{summary_row},0)")
            avg_gain_loss.border = border
            avg_gain_loss.number_format = '0.00%'
            avg_gain_loss.font = Font(bold=True)
            
            # Update cash percentage if cash row was added
            if cash_pct_row is not None:
                # Cash % = Cash Amount / Total Portfolio Value (including cash)
                cash_pct_formula = f"=F{cash_pct_row}/F${summary_row}"
                sheet.cell(row=cash_pct_row, column=9, value=cash_pct_formula)
            
            # Apply conditional formatting to summary row
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Market Value Total (positive/negative coloring)
            sheet.conditional_formatting.add(f'F{summary_row}:F{summary_row}', 
                                           CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))
            sheet.conditional_formatting.add(f'F{summary_row}:F{summary_row}', 
                                           CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
            
            # Unrealized Gain/Loss Total (positive/negative coloring)
            sheet.conditional_formatting.add(f'G{summary_row}:G{summary_row}', 
                                           CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))
            sheet.conditional_formatting.add(f'G{summary_row}:G{summary_row}', 
                                           CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
            
            # Average % Gain/Loss (positive/negative coloring)
            sheet.conditional_formatting.add(f'H{summary_row}:H{summary_row}', 
                                           CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))
            sheet.conditional_formatting.add(f'H{summary_row}:H{summary_row}', 
                                           CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
            
            current_row += 1
            
            # Add pie charts to the right of the section
            chart_start_col = "M"  # Column M (13)
            chart_row_offset = start_row + 1  # Start charts at header level
            chart_start_col_num = column_index_from_string(chart_start_col)  # M = 13
            
            # Second chart column - large offset to prevent overlap with first chart
            chart2_col_num = chart_start_col_num + 32  # M(13) + 32 = 45 (Column AS)
            
            # Helper data columns starting at BA (53) to avoid overlap with charts
            # Symbol distribution helper: columns BA-BB (53-54)
            # Sector distribution helper: columns BE-BF (57-58)
            symbol_helper_col = 53  # Column BA
            sector_helper_col = 57  # Column BE (with gap to avoid overlap)
            
            # Calculate helper data row based on section's start_row to ensure unique rows per section
            # Use the section's data rows as reference to keep helper data aligned with its section
            helper_start_row = start_row + 1  # Start helper data at same row as section header
            
            # Get portfolio percentages and sort in descending order
            # Only include data from THIS section's rows (data_start to data_end)
            portfolio_data_for_sort = []
            for row_num in range(data_start, data_end + 1):
                symbol = sheet.cell(row=row_num, column=1).value
                shares = sheet.cell(row=row_num, column=3).value or 0
                current_price = sheet.cell(row=row_num, column=5).value or 0
                market_value = shares * current_price if (shares and current_price) else 0
                portfolio_data_for_sort.append((symbol, f"=I{row_num}", row_num, market_value))
            
            # Sort by market value in descending order (proxy for portfolio percentage)
            portfolio_data_for_sort.sort(key=lambda x: x[3], reverse=True)
            
            # Create sorted helper data for portfolio distribution chart
            sorted_helper_row = helper_start_row
            sheet.cell(row=sorted_helper_row, column=symbol_helper_col, value="Sorted Symbols").font = Font(bold=True)
            sheet.cell(row=sorted_helper_row, column=symbol_helper_col + 1, value="Sorted % Portfolio").font = Font(bold=True)
            sorted_helper_row += 1
            
            # Create properly sorted helper data
            sorted_data_start = sorted_helper_row
            for symbol, pct_ref, original_row, _ in portfolio_data_for_sort:
                sheet.cell(row=sorted_helper_row, column=symbol_helper_col, value=symbol)
                sheet.cell(row=sorted_helper_row, column=symbol_helper_col + 1, value=pct_ref)
                sorted_helper_row += 1
            sorted_data_end = sorted_helper_row - 1
            
            # Chart 1: Portfolio Distribution by Symbol (sorted descending)
            pie_symbols = PieChart()
            pie_symbols.title = f"{section_title} - Portfolio Distribution"
            pie_symbols.width = 12
            pie_symbols.height = 10
            
            # Use sorted helper data
            if sorted_data_end >= sorted_data_start:
                symbols_sorted_range = Reference(sheet, min_col=symbol_helper_col, min_row=sorted_data_start, max_row=sorted_data_end)
                portfolio_sorted_range = Reference(sheet, min_col=symbol_helper_col + 1, min_row=sorted_data_start, max_row=sorted_data_end)
                
                pie_symbols.add_data(portfolio_sorted_range)
                pie_symbols.set_categories(symbols_sorted_range)
            else:
                # Fallback to original data if helper data creation failed
                symbols_range = Reference(sheet, min_col=1, min_row=data_start, max_row=data_end)
                portfolio_pct_range = Reference(sheet, min_col=9, min_row=data_start, max_row=data_end)
                pie_symbols.add_data(portfolio_pct_range)
                pie_symbols.set_categories(symbols_range)
            
            # Add first chart
            sheet.add_chart(pie_symbols, f"{chart_start_col}{chart_row_offset}")
            
            # Chart 2: Sector Distribution with aggregation
            pie_sectors = PieChart()
            pie_sectors.title = f"{section_title} - Sector Distribution"
            pie_sectors.width = 12
            pie_sectors.height = 10
            
            # Create aggregated sector data with sorting
            # Only include data from THIS section's rows (data_start to data_end)
            sector_aggregation = {}
            sector_market_values = {}
            
            for row_num in range(data_start, data_end + 1):
                sector = sheet.cell(row=row_num, column=10).value
                shares = sheet.cell(row=row_num, column=3).value or 0
                current_price = sheet.cell(row=row_num, column=5).value or 0
                market_value = shares * current_price if (shares and current_price) else 0
                
                if sector not in sector_aggregation:
                    sector_aggregation[sector] = []
                    sector_market_values[sector] = 0
                    
                sector_aggregation[sector].append(row_num)
                sector_market_values[sector] += market_value
            
            # Sort sectors by their total market value in descending order
            sorted_sectors = sorted(sector_aggregation.keys(), 
                                  key=lambda s: sector_market_values[s], reverse=True)
            
            # Create helper data for sector aggregation (sorted by value)
            sector_helper_row = helper_start_row
            sheet.cell(row=sector_helper_row, column=sector_helper_col, value="Distinct Sectors").font = Font(bold=True)
            sheet.cell(row=sector_helper_row, column=sector_helper_col + 1, value="Aggregated % Portfolio").font = Font(bold=True)
            sector_helper_row += 1
            
            sector_data_start = sector_helper_row
            for sector in sorted_sectors:
                row_nums = sector_aggregation[sector]
                # Sector name
                sheet.cell(row=sector_helper_row, column=sector_helper_col, value=sector)
                
                # Sum of % portfolio for this sector (using only rows from THIS section)
                sum_formula = "+".join([f"I{row_num}" for row_num in row_nums])
                sheet.cell(row=sector_helper_row, column=sector_helper_col + 1, value=f"={sum_formula}")
                
                sector_helper_row += 1
            
            sector_data_end = sector_helper_row - 1
            
            # Use aggregated data for sector pie chart
            if sector_data_end >= sector_data_start:
                sectors_agg_range = Reference(sheet, min_col=sector_helper_col, min_row=sector_data_start, max_row=sector_data_end)
                sectors_values_range = Reference(sheet, min_col=sector_helper_col + 1, min_row=sector_data_start, max_row=sector_data_end)
                
                pie_sectors.add_data(sectors_values_range)
                pie_sectors.set_categories(sectors_agg_range)
            
            # Add second chart (using pre-calculated offset to avoid overlap)
            chart2_col = get_column_letter(chart2_col_num)
            sheet.add_chart(pie_sectors, f"{chart2_col}{chart_row_offset}")
            
        # Add spacing between sections (extra space for charts)
        current_row += 15  # Extra space for charts
        
        # Return summary row location if summary was created
        summary_row_location = summary_row if holdings_rows else None
        
        return current_row, len(holdings_rows), data_rows, summary_row_location
    
    current_row = 1
    all_holding_rows = []
    section_summary_rows = []  # Track summary rows for global calculation
    
    # Create sections for each account
    for account_name, account_data in portfolio_data['accounts'].items():
        # Get cash information from account data
        account_cash = account_data.get('cash', 0)
        account_currency = account_data.get('currency', '$')
        
        # Create section if account has holdings OR cash > 0
        if account_data['holdings'] or account_cash > 0:
            section_title = f"ACCOUNT: {account_name.upper()}"
            
            current_row, num_rows, section_data_rows, summary_row_loc = create_holdings_section(
                current_row, section_title, account_data['holdings'], 
                is_global=False, section_summary_rows=None,
                cash_amount=account_cash, cash_currency=account_currency)
            all_holding_rows.extend(section_data_rows)
            # Track the summary row for global calculation
            if summary_row_loc:
                section_summary_rows.append(summary_row_loc)
    
    # Calculate total cash for global sections
    total_cash_usd = 0
    total_cash_eur = 0
    exchange_rates = portfolio_data.get('exchange_rates', {})
    
    for account_name, account_data in portfolio_data['accounts'].items():
        account_cash = account_data.get('cash', 0)
        account_currency = account_data.get('currency', '$')
        
        if account_cash > 0:
            # Convert to USD
            cash_usd = convert_to_usd(account_cash, account_currency, exchange_rates)
            total_cash_usd += cash_usd
            
            # Convert to EUR
            cash_eur = convert_to_eur(account_cash, account_currency, exchange_rates)
            total_cash_eur += cash_eur
    
    # Create global portfolio section with USD conversion
    if portfolio_data['holdings']:
        section_title = "GLOBAL PORTFOLIO (USD)"
        
        # Convert all holdings to USD for global portfolio section
        usd_holdings = create_global_holdings_with_usd_conversion(portfolio_data['holdings'], exchange_rates)
        
        print(f"🌍 Global portfolio: Converting {len(usd_holdings)} holdings to USD")
        for symbol, data in usd_holdings.items():
            original_currency = data.get('original_currency', '$')
            if original_currency != '$':
                # Map currency symbol to code for rate lookup
                currency_code_map = {
                    '€': 'EUR',
                    'RON': 'RON', 
                    '£': 'GBP',
                    '¥': 'JPY'
                }
                currency_code = currency_code_map.get(original_currency, original_currency)
                rate = exchange_rates.get(currency_code, 1.0)
                print(f"   {symbol}: {original_currency} → USD (rate: {rate:.4f})")
        
        current_row, num_rows, section_data_rows, summary_row_loc = create_holdings_section(
            current_row, section_title, usd_holdings, is_global=True, 
            section_summary_rows=section_summary_rows,
            cash_amount=total_cash_usd, cash_currency='$')
        all_holding_rows.extend(section_data_rows)
        
    # Create global portfolio section with EUR conversion
    if portfolio_data['holdings']:
        section_title = "GLOBAL PORTFOLIO (EUR)"
        
        # Convert all holdings to EUR for global portfolio section
        exchange_rates = portfolio_data.get('exchange_rates', {})
        eur_holdings = create_global_holdings_with_eur_conversion(portfolio_data['holdings'], exchange_rates)
        
        print(f"🇪🇺 Global portfolio: Converting {len(eur_holdings)} holdings to EUR")
        for symbol, data in eur_holdings.items():
            original_currency = data.get('original_currency', '$')
            if original_currency != '€':
                # Map currency symbol to code for rate lookup
                currency_code_map = {
                    '$': 'USD',
                    'RON': 'RON', 
                    '£': 'GBP',
                    '¥': 'JPY'
                }
                currency_code = currency_code_map.get(original_currency, original_currency)
                eur_rates = exchange_rates.get('EUR_RATES', {})
                rate = eur_rates.get(currency_code, 1.0)
                print(f"   {symbol}: {original_currency} → EUR (rate: {rate:.4f})")
        
        current_row, num_rows, section_data_rows, summary_row_loc = create_holdings_section(
            current_row, section_title, eur_holdings, is_global=True, 
            section_summary_rows=section_summary_rows,
            cash_amount=total_cash_eur, cash_currency='€')
    all_holding_rows.extend(section_data_rows)
    
    # Add conditional formatting for gains/losses across all sections
    if all_holding_rows:
        # Colors for gain/loss columns
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Colors for % of Portfolio column
        portfolio_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # > 20%
        portfolio_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 10-20%
        portfolio_orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # 5-10%
        portfolio_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # 0-5%
        
        for row in all_holding_rows:
            # Apply to gain/loss columns (G and H)
            sheet.conditional_formatting.add(f'G{row}:G{row}', 
                                           CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))
            sheet.conditional_formatting.add(f'G{row}:G{row}', 
                                           CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
            sheet.conditional_formatting.add(f'H{row}:H{row}', 
                                           CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))
            sheet.conditional_formatting.add(f'H{row}:H{row}', 
                                           CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
            
            # Apply to % of Portfolio column (I) with specific thresholds
            # > 20% - Green
            sheet.conditional_formatting.add(f'I{row}:I{row}', 
                                           CellIsRule(operator='greaterThan', formula=['0.2'], fill=portfolio_green))
            
            # 10% <= x < 20% - Yellow  
            sheet.conditional_formatting.add(f'I{row}:I{row}', 
                                           CellIsRule(operator='between', formula=['0.1', '0.199'], fill=portfolio_yellow))
            
            # 5% <= x < 10% - Orange
            sheet.conditional_formatting.add(f'I{row}:I{row}', 
                                           CellIsRule(operator='between', formula=['0.05', '0.099'], fill=portfolio_orange))
            
            # 0% <= x < 5% - Red
            sheet.conditional_formatting.add(f'I{row}:I{row}', 
                                           CellIsRule(operator='between', formula=['0', '0.049'], fill=portfolio_red))
    
    # Auto-adjust column widths
    auto_adjust_column_widths(sheet)

def create_transactions_history(sheet, header_font, header_fill, border, portfolio_data):
    """Create transaction history"""
    headers = ["Date", "Type", "Symbol", "Company Name", "Shares",
               "Price", "Total Amount", "Fees", "Net Amount", "Notes"]
    
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    
    transactions = portfolio_data['transactions']

    # Sort transactions by date (newest first)
    try:
        transactions.sort(key=lambda x: datetime.strptime(x['date'], "%d-%m-%Y"), reverse=True)
    except:
        pass  # If date parsing fails, use original order
    
    # Populate the sheet with transactions
    for idx, transaction in enumerate(transactions, 2):
        row = idx
        trans_type = transaction['type'].upper()
        shares = transaction['shares']
        price = transaction['price']
        fees = transaction['fee']
        notes = transaction.get('note', '')
        currency = transaction.get('currency', '$')  # Get currency from transaction
        currency_format = get_currency_format(currency)
        
        # Convert date format if needed
        try:
            if '-' in transaction['date'] and len(transaction['date'].split('-')[0]) == 2:
                # Convert from DD-MM-YYYY to YYYY-MM-DD
                date_parts = transaction['date'].split('-')
                formatted_date = f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
            else:
                formatted_date = transaction['date']
        except:
            formatted_date = transaction['date']
        
        # Fill in the basic data first
        basic_data = [formatted_date, trans_type, transaction['symbol'], transaction['name'], shares, price]
        
        for col, value in enumerate(basic_data, 1):
            cell = sheet.cell(row=row, column=col, value=value)
            cell.border = border
            if col == 1:
                cell.font = Font(bold=True)
            elif col == 5 and type(shares) == float:  # Shares
                cell.number_format = '#,##0.0000'
            elif col == 6:  # Price column
                cell.number_format = currency_format
        
        # Total Amount column (F) - use formula
        total_amount_cell = sheet.cell(row=row, column=7, value=f"=E{row}*F{row}")
        total_amount_cell.border = border
        total_amount_cell.number_format = currency_format
        
        # Fees column (G)
        fees_cell = sheet.cell(row=row, column=8, value=fees)
        fees_cell.border = border
        fees_cell.number_format = currency_format
        
        # Net Amount column (H) - use formula
        net_amount_cell = sheet.cell(row=row, column=9, value=f"=G{row}-H{row}")
        net_amount_cell.border = border
        net_amount_cell.number_format = currency_format
        
        # Notes column (I)
        notes_cell = sheet.cell(row=row, column=10, value=notes)
        notes_cell.border = border
    
    # Auto-adjust column widths
    auto_adjust_column_widths(sheet)

def create_dividend_tracker(sheet, header_font, header_fill, border, portfolio_data):
    """Create dividend tracking sheet with formulas and charts"""
    headers = ["Symbol", "Company Name", "Shares", "Current Price", "Annual Dividend", "Quarterly Dividend", 
               "Yield %", "Ex-Date", "Pay Date", "YTD Received", "Projected Annual"]
    
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    
    # Get dividend data from portfolio holdings - only stocks (exclude ETF, Crypto, Cash)
    stock_holdings = []
    real_time_data = portfolio_data.get('real_time_data', {})
    
    # Stock asset classes to include
    stock_asset_classes = ["US Stocks", "Romanian Stocks", "International Stocks"]
    
    for symbol, holding in portfolio_data['holdings'].items():
        asset_class = holding.get('asset_class', 'Unknown')
        
        # Only include stock holdings (exclude ETF, Crypto, Cash)
        if asset_class in stock_asset_classes:
            # Get company name and shares from holdings
            company_name = holding.get('company_name', holding.get('name', symbol))
            shares = holding['total_shares']
            currency = holding.get('currency', '$')
            
            # Get dividend data from yfinance if available
            if symbol in real_time_data:
                rt_data = real_time_data[symbol]
                current_price = rt_data.get('current_price', holding['weighted_avg_cost'])
                dividend_yield = rt_data.get('dividend_yield', 0)
                
                # Handle dividend yield (could be None or 'N/A')
                try:
                    dividend_yield = float(dividend_yield) if dividend_yield else 0
                except (ValueError, TypeError):
                    dividend_yield = 0
                
                # Calculate annual dividend per share from yield and current price
                annual_dividend = current_price * dividend_yield if dividend_yield else 0
            else:
                # No real-time data available
                current_price = holding['weighted_avg_cost']
                dividend_yield = 0
                annual_dividend = 0
            
            stock_holdings.append({
                'symbol': symbol,
                'company_name': company_name,
                'shares': shares,
                'current_price': current_price,
                'annual_dividend': annual_dividend,
                'dividend_yield': dividend_yield,
                'currency': currency
            })
    
    # Sort by symbol for consistent ordering
    stock_holdings.sort(key=lambda x: x['symbol'])
    
    # Populate dividend data
    for row, stock in enumerate(stock_holdings, 2):
        currency_format = get_currency_format(stock['currency'])
        
        # Symbol
        symbol_cell = sheet.cell(row=row, column=1, value=stock['symbol'])
        symbol_cell.border = border
        symbol_cell.font = Font(bold=True)

        # Company Name
        company_cell = sheet.cell(row=row, column=2, value=stock['company_name'])
        company_cell.border = border
        
        # Shares
        shares_cell = sheet.cell(row=row, column=3, value=stock['shares'])
        shares_cell.border = border
        shares_cell.number_format = '#,##0.0000'
        
        # Current Price
        price_cell = sheet.cell(row=row, column=4, value=stock['current_price'])
        price_cell.border = border
        price_cell.number_format = currency_format
        
        # Annual dividend per share
        annual_div_cell = sheet.cell(row=row, column=5, value=stock['annual_dividend'])
        annual_div_cell.border = border
        annual_div_cell.number_format = currency_format
        
        # Quarterly dividend = Annual / 4
        quarterly_cell = sheet.cell(row=row, column=6, value=f"=E{row}/4")
        quarterly_cell.border = border
        quarterly_cell.number_format = currency_format
        
        # Dividend Yield % (from yfinance data)
        yield_cell = sheet.cell(row=row, column=7, value=stock['dividend_yield'])
        yield_cell.border = border
        yield_cell.number_format = '0.00%'
        
        # Ex-Date and Pay-Date - not available from yfinance basic data
        sheet.cell(row=row, column=8, value="N/A").border = border
        sheet.cell(row=row, column=9, value="N/A").border = border
        
        # YTD Received = Annual Dividend * Shares * 0.75 (assuming 3/4 of year passed)
        ytd_cell = sheet.cell(row=row, column=10, value=f"=E{row}*C{row}*0.75")
        ytd_cell.border = border
        ytd_cell.number_format = currency_format
        
        # Projected Annual = Annual Dividend * Shares
        projected_cell = sheet.cell(row=row, column=11, value=f"=E{row}*C{row}")
        projected_cell.border = border
        projected_cell.number_format = currency_format
    
    # Only add charts if there are stock holdings
    if stock_holdings:
        max_row = len(stock_holdings) + 1
        
        # Position charts below the dividend data
        chart_start_row = max_row + 3
        
        # Add Dividend Yield Bar Chart
        bar_chart = BarChart()
        bar_chart.title = "Dividend Yields by Position"
        bar_chart.style = 10
        bar_chart.y_axis.title = "Yield %"
        bar_chart.x_axis.title = "Stock"
        bar_chart.width = 15
        bar_chart.height = 10
        
        # Data for the chart
        data = Reference(sheet, min_col=7, min_row=1, max_row=max_row)
        cats = Reference(sheet, min_col=1, min_row=2, max_row=max_row)
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(cats)
        
        # Add the chart to the sheet
        sheet.add_chart(bar_chart, f"A{chart_start_row}")
        
        # Add Dividend Income Pie Chart (only for stocks with dividends)
        stocks_with_dividends = [s for s in stock_holdings if s['annual_dividend'] > 0]
        if stocks_with_dividends:
            pie = PieChart()
            pie.title = "Projected Annual Dividend Income"
            pie.width = 12
            pie.height = 10
            
            # Create helper data for pie chart (only dividend-paying stocks)
            helper_row = chart_start_row + 16  # Position helper data below first chart
            sheet.cell(row=helper_row, column=1, value="Symbol").font = Font(bold=True)
            sheet.cell(row=helper_row, column=2, value="Annual Income").font = Font(bold=True)
            helper_row += 1
            
            helper_start = helper_row
            for stock in stocks_with_dividends:
                sheet.cell(row=helper_row, column=1, value=stock['symbol'])
                # Calculate projected annual income
                annual_income = stock['annual_dividend'] * stock['shares']
                sheet.cell(row=helper_row, column=2, value=annual_income)
                helper_row += 1
            helper_end = helper_row - 1
            
            # Add pie chart data
            labels = Reference(sheet, min_col=1, min_row=helper_start, max_row=helper_end)
            pie_data = Reference(sheet, min_col=2, min_row=helper_start, max_row=helper_end)
            pie.add_data(pie_data)
            pie.set_categories(labels)
            
            # Add the pie chart below the bar chart
            pie_chart_row = chart_start_row + 16  # Position pie chart below bar chart
            sheet.add_chart(pie, f"A{pie_chart_row}")
    
    # Auto-adjust column widths
    auto_adjust_column_widths(sheet)

def get_historical_price(symbol, target_date):
    """Get historical price for a symbol at a specific date using yfinance"""
    if not YFINANCE_AVAILABLE:
        return None
    
    try:
        stock = yf.Ticker(symbol)
        # Get data around the target date (a few days before and after to ensure we have data)
        start_date = target_date - timedelta(days=10)
        end_date = target_date + timedelta(days=5)
        
        hist = stock.history(start=start_date, end=end_date)
        
        if hist.empty:
            return None
        
        # Find the closest date that is <= target_date
        hist.index = hist.index.tz_localize(None)  # Remove timezone for comparison
        valid_dates = hist.index[hist.index <= target_date]
        
        if len(valid_dates) == 0:
            # If no dates before target, use the first available date
            return hist['Close'].iloc[0]
        
        # Get the last valid date (closest to target)
        closest_date = valid_dates[-1]
        return hist.loc[closest_date, 'Close']
        
    except Exception as e:
        print(f"  ⚠️  Could not get historical price for {symbol}: {e}")
        return None

def get_sp500_price(target_date):
    """Get S&P 500 price at a specific date"""
    return get_historical_price("^GSPC", target_date)

def compute_portfolio_performance(portfolio_data, target_month, target_year, exchange_rates=None):
    """
    Compute portfolio performance up to a given date.
    
    Args:
        portfolio_data: The full portfolio data dictionary with all transactions
        target_month: The target month (1-12)
        target_year: The target year (e.g., 2025)
        exchange_rates: Exchange rates for currency conversion (optional)
    
    Returns:
        Dictionary containing:
        - portfolio_value: Total portfolio value in USD at the target date
        - total_invested: Total amount invested up to that date
        - total_gain_loss: Total gain/loss
        - holdings: Dictionary of holdings at that date
    """
    # Determine the target date (end of month, or current date if current month)
    current_date = datetime.now()
    
    if target_year == current_date.year and target_month == current_date.month:
        # Current month - use current date
        target_date = current_date
        use_current_prices = True
    else:
        # Historical month - use end of month
        last_day = monthrange(target_year, target_month)[1]
        target_date = datetime(target_year, target_month, last_day)
        use_current_prices = False
    
    # Get exchange rates if not provided
    if exchange_rates is None:
        exchange_rates = get_exchange_rates()
    
    # Collect all transactions up to target date
    all_transactions = []
    
    for account in portfolio_data.get('accounts', {}).values():
        for transaction in account.get('transactions', []):
            try:
                trans_date = datetime.strptime(transaction['date'], "%d-%m-%Y")
            except:
                continue
            
            if trans_date <= target_date:
                all_transactions.append(transaction)
    
    # Sort transactions by date
    try:
        all_transactions.sort(key=lambda x: datetime.strptime(x['date'], "%d-%m-%Y"))
    except:
        pass
    
    # Calculate holdings at target date using FIFO
    holdings = defaultdict(lambda: {
        'name': '',
        'lots': deque(),
        'total_shares': 0,
        'weighted_avg_cost': 0,
        'total_cost': 0,
        'currency': '$'
    })
    
    for transaction in all_transactions:
        symbol = transaction['symbol']
        shares = transaction['shares']
        price = transaction['price']
        trans_type = transaction['type'].lower()
        date = transaction['date']
        currency = transaction.get('currency', '$')
        
        if trans_type == 'buy':
            holdings[symbol]['lots'].append((shares, price, date))
            holdings[symbol]['name'] = transaction['name']
            holdings[symbol]['currency'] = currency
            
        elif trans_type == 'sell':
            remaining_to_sell = shares
            
            while remaining_to_sell > 0 and holdings[symbol]['lots']:
                lot_shares, lot_price, lot_date = holdings[symbol]['lots'][0]
                
                if lot_shares <= remaining_to_sell:
                    remaining_to_sell -= lot_shares
                    holdings[symbol]['lots'].popleft()
                else:
                    new_lot_shares = lot_shares - remaining_to_sell
                    holdings[symbol]['lots'][0] = (new_lot_shares, lot_price, lot_date)
                    remaining_to_sell = 0
        
        # Recalculate totals for this symbol
        total_shares = 0
        total_cost = 0
        for s, p, _ in holdings[symbol]['lots']:
            total_shares += s
            total_cost += s * p
        
        holdings[symbol]['total_shares'] = total_shares
        holdings[symbol]['total_cost'] = total_cost
        holdings[symbol]['weighted_avg_cost'] = total_cost / total_shares if total_shares > 0 else 0
    
    # Remove zero holdings
    holdings = {k: v for k, v in holdings.items() if v['total_shares'] > 0}
    
    # Calculate portfolio value at target date
    portfolio_value = 0
    
    # Price cache for this calculation
    price_cache = {}
    
    # Track % Gain/Loss for each holding to calculate the average
    holdings_gain_loss_list = []
    
    for symbol, holding in holdings.items():
        shares = holding['total_shares']
        currency = holding['currency']
        avg_cost = holding['weighted_avg_cost']
        current_price = None
        
        if use_current_prices:
            # Use real-time prices from portfolio_data if available
            real_time_data = portfolio_data.get('real_time_data', {})
            if symbol in real_time_data:
                current_price = real_time_data[symbol].get('current_price')
                if current_price:
                    price_cache[symbol] = current_price
        
        # Try to get historical price if not yet found
        if current_price is None:
            if symbol not in price_cache:
                hist_price = get_historical_price(symbol, target_date)
                if hist_price is not None:
                    price_cache[symbol] = hist_price
                else:
                    # Fallback to weighted average cost
                    price_cache[symbol] = holding['weighted_avg_cost']
            current_price = price_cache[symbol]
        
        # Calculate market value
        price_usd = convert_to_usd(current_price, currency, exchange_rates)
        market_value = shares * price_usd
        portfolio_value += market_value
        
        # Calculate % Gain/Loss for this holding
        if avg_cost > 0:
            pct_gain_loss = (current_price - avg_cost) / avg_cost
            holdings_gain_loss_list.append(pct_gain_loss)
    
    # Calculate average % Gain/Loss across all holdings
    if holdings_gain_loss_list:
        avg_pct_gain_loss = sum(holdings_gain_loss_list) / len(holdings_gain_loss_list)
    else:
        avg_pct_gain_loss = 0

    return {
        'date': target_date,
        'portfolio_value': portfolio_value,
        'holdings_count': len(holdings),
        'avg_pct_gain_loss': avg_pct_gain_loss
    }

def get_month_range_from_transactions(portfolio_data):
    """Get the range of months from oldest transaction to current date"""
    # Find oldest transaction date
    oldest_date = None
    
    for account in portfolio_data.get('accounts', {}).values():
        for transaction in account.get('transactions', []):
            try:
                trans_date = datetime.strptime(transaction['date'], "%d-%m-%Y")
                if oldest_date is None or trans_date < oldest_date:
                    oldest_date = trans_date
            except:
                continue
    
    if oldest_date is None:
        return []
    
    # Generate list of (year, month) tuples from oldest to current
    current_date = datetime.now()
    months = []
    
    year = oldest_date.year
    month = oldest_date.month
    
    while (year < current_date.year) or (year == current_date.year and month <= current_date.month):
        months.append((year, month))
        
        month += 1
        if month > 12:
            month = 1
            year += 1
    
    return months

def create_performance_analysis(sheet, header_font, header_fill, border, portfolio_data):
    """Create performance analysis sheet with monthly portfolio performance data"""
    
    headers = ["Month", "Portfolio Value", "Monthly Return %", "S&P 500 Return %", 
               "Outperformance", "Cumulative Return %", "Benchmark Cumulative %", "Alpha"]
    
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    
    # Get month range from transactions
    months = get_month_range_from_transactions(portfolio_data)
    
    if not months:
        print("  ⚠️  No transactions found for performance analysis")
        auto_adjust_column_widths(sheet)
        return
    
    print(f"\n📈 Computing monthly portfolio performance for {len(months)} months...")
    print(f"   From: {months[0][0]}-{months[0][1]:02d} to {months[-1][0]}-{months[-1][1]:02d}")
    
    # Get exchange rates
    exchange_rates = portfolio_data.get('exchange_rates', get_exchange_rates())
    
    # Track for cumulative calculations
    previous_avg_pct_gain_loss = None
    previous_sp500_price = None
    cumulative_factor = 1.0  # Compounding factor for portfolio: (1+r1)*(1+r2)*...
    benchmark_factor = 1.0  # Compounding factor for S&P 500
    
    # Cache for S&P 500 prices
    sp500_cache = {}
    
    current_row = 2
    
    for idx, (year, month) in enumerate(months):
        month_name = datetime(year, month, 1).strftime("%b %Y")
        print(f"  [{idx+1}/{len(months)}] Processing {month_name}...")
        
        # Compute portfolio performance for this month
        perf = compute_portfolio_performance(portfolio_data, month, year, exchange_rates)
        portfolio_value = perf['portfolio_value']
        avg_pct_gain_loss = perf['avg_pct_gain_loss']
        
        # Get end of month date for S&P 500
        if year == datetime.now().year and month == datetime.now().month:
            target_date = datetime.now()
        else:
            last_day = monthrange(year, month)[1]
            target_date = datetime(year, month, last_day)
        
        # Get S&P 500 price
        month_key = f"{year}-{month:02d}"
        if month_key not in sp500_cache:
            sp500_price = get_sp500_price(target_date)
            sp500_cache[month_key] = sp500_price
        else:
            sp500_price = sp500_cache[month_key]
        
        # Calculate monthly return as difference between current and previous month's avg % Gain/Loss
        if previous_avg_pct_gain_loss is not None:
            monthly_return = avg_pct_gain_loss - previous_avg_pct_gain_loss
        else:
            monthly_return = avg_pct_gain_loss  # First month: use the avg % Gain/Loss itself
        
        # Calculate S&P 500 monthly return
        if previous_sp500_price and previous_sp500_price > 0 and sp500_price:
            sp500_monthly_return = (sp500_price - previous_sp500_price) / previous_sp500_price
        else:
            sp500_monthly_return = 0
        
        # Calculate outperformance
        outperformance = monthly_return - sp500_monthly_return
        
        # Calculate cumulative returns with compounding effect
        # Formula: (1 + r1) * (1 + r2) * ... * (1 + rn) - 1
        cumulative_factor *= (1 + monthly_return)
        benchmark_factor *= (1 + sp500_monthly_return)
        
        cumulative_return = cumulative_factor - 1
        benchmark_cumulative = benchmark_factor - 1
        
        # Calculate alpha (cumulative outperformance)
        alpha = cumulative_return - benchmark_cumulative
        
        # Write to sheet
        # Month
        month_cell = sheet.cell(row=current_row, column=1, value=month_name)
        month_cell.border = border
        month_cell.font = Font(bold=True)
        
        # Portfolio Value
        value_cell = sheet.cell(row=current_row, column=2, value=portfolio_value)
        value_cell.border = border
        value_cell.number_format = '$#,##0'
        
        # Monthly Return %
        monthly_return_cell = sheet.cell(row=current_row, column=3, value=monthly_return)
        monthly_return_cell.border = border
        monthly_return_cell.number_format = '0.00%'
        
        # S&P 500 Return %
        sp500_cell = sheet.cell(row=current_row, column=4, value=sp500_monthly_return)
        sp500_cell.border = border
        sp500_cell.number_format = '0.00%'
        
        # Outperformance
        outperf_cell = sheet.cell(row=current_row, column=5, value=outperformance)
        outperf_cell.border = border
        outperf_cell.number_format = '0.00%'
        
        # Cumulative Return %
        cum_return_cell = sheet.cell(row=current_row, column=6, value=cumulative_return)
        cum_return_cell.border = border
        cum_return_cell.number_format = '0.00%'
        
        # Benchmark Cumulative %
        bench_cum_cell = sheet.cell(row=current_row, column=7, value=benchmark_cumulative)
        bench_cum_cell.border = border
        bench_cum_cell.number_format = '0.00%'
        
        # Alpha
        alpha_cell = sheet.cell(row=current_row, column=8, value=alpha)
        alpha_cell.border = border
        alpha_cell.number_format = '0.00%'
        
        # Update tracking values for next iteration
        previous_avg_pct_gain_loss = avg_pct_gain_loss
        if sp500_price:
            previous_sp500_price = sp500_price
        
        current_row += 1
    
    # Apply conditional formatting for returns and alpha
    data_end_row = current_row - 1
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Monthly Return (Column C)
    sheet.conditional_formatting.add(f'C2:C{data_end_row}', 
                                   CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))
    sheet.conditional_formatting.add(f'C2:C{data_end_row}', 
                                   CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
    
    # Outperformance (Column E)
    sheet.conditional_formatting.add(f'E2:E{data_end_row}', 
                                   CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))
    sheet.conditional_formatting.add(f'E2:E{data_end_row}', 
                                   CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
    
    # Alpha (Column H)
    sheet.conditional_formatting.add(f'H2:H{data_end_row}', 
                                   CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))
    sheet.conditional_formatting.add(f'H2:H{data_end_row}', 
                                   CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
    
    # Add charts below the data
    if data_end_row > 1:
        chart_start_row = data_end_row + 3
        
        # Portfolio Value Chart (Line Chart)
        line_chart = LineChart()
        line_chart.title = "Portfolio Value Over Time"
        line_chart.style = 10
        line_chart.y_axis.title = "Value (USD)"
        line_chart.x_axis.title = "Month"
        line_chart.width = 18
        line_chart.height = 10
        
        data = Reference(sheet, min_col=2, min_row=1, max_row=data_end_row)
        cats = Reference(sheet, min_col=1, min_row=2, max_row=data_end_row)
        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(cats)
        
        sheet.add_chart(line_chart, f"A{chart_start_row}")
        
        # Cumulative Returns Comparison Chart
        comparison_chart = LineChart()
        comparison_chart.title = "Portfolio vs S&P 500 Cumulative Returns"
        comparison_chart.style = 10
        comparison_chart.y_axis.title = "Return %"
        comparison_chart.x_axis.title = "Month"
        comparison_chart.width = 18
        comparison_chart.height = 10
        
        # Add both cumulative return columns
        data_portfolio = Reference(sheet, min_col=6, min_row=1, max_row=data_end_row)
        data_benchmark = Reference(sheet, min_col=7, min_row=1, max_row=data_end_row)
        comparison_chart.add_data(data_portfolio, titles_from_data=True)
        comparison_chart.add_data(data_benchmark, titles_from_data=True)
        comparison_chart.set_categories(cats)
        
        sheet.add_chart(comparison_chart, f"J{chart_start_row}")
    
    print(f"  ✅ Performance analysis complete - {data_end_row - 1} months processed")
    
    # Auto-adjust column widths
    auto_adjust_column_widths(sheet)

def create_portfolio_overview(sheet, header_font, header_fill, border, portfolio_data=None):
    """Create portfolio overview with asset allocation analysis based on Holdings sheet"""
    headers = ["Asset Class", "Current %", "Target %", "Current Value", 
               "Target Value", "Difference", "Action Required"]
    
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    
    # Get distinct asset classes from portfolio data
    asset_classes_in_portfolio = set()
    for holding in portfolio_data['holdings'].values():
        asset_class = holding.get('asset_class', 'US Stocks')
        asset_classes_in_portfolio.add(asset_class)
    
    # Get target allocations from portfolio data
    target_allocations = portfolio_data.get('target_asset_class_distribution', {})
    if target_allocations:
        print(f"📊 Using target allocations from portfolio.json: {target_allocations}")
    else:
        print(f"⚠️  No target allocations found in portfolio.json")
    
    # Calculate current values by asset class from the Global Portfolio (USD) data
    asset_class_values = {}
    total_portfolio_value = 0
    
    # Use the Global Portfolio (USD) holdings data (converted to USD)
    exchange_rates = portfolio_data.get('exchange_rates', {})
    global_usd_holdings = create_global_holdings_with_usd_conversion(portfolio_data['holdings'], exchange_rates)
    
    for symbol, holding in global_usd_holdings.items():
        asset_class = holding.get('asset_class', 'US Stocks')
        # Calculate market value in USD
        current_price = holding.get('current_price', holding['weighted_avg_cost'])
        market_value = holding['total_shares'] * current_price
        
        if asset_class not in asset_class_values:
            asset_class_values[asset_class] = 0
        asset_class_values[asset_class] += market_value
        total_portfolio_value += market_value
    
    # Add cash as an asset class (converted to USD)
    total_cash_usd = 0
    for account_name, account_data in portfolio_data['accounts'].items():
        account_cash = account_data.get('cash', 0)
        account_currency = account_data.get('currency', '$')
        
        if account_cash > 0:
            # Convert to USD
            cash_usd = convert_to_usd(account_cash, account_currency, exchange_rates)
            total_cash_usd += cash_usd
    
    if total_cash_usd > 0:
        asset_class_values['Cash'] = total_cash_usd
        total_portfolio_value += total_cash_usd
        asset_classes_in_portfolio.add('Cash')
    
    # Sort asset classes for consistent ordering (cash will be included)
    sorted_asset_classes = sorted(asset_classes_in_portfolio)
    
    # Create rows for each asset class
    current_row = 2
    data_end_row = current_row + len(sorted_asset_classes) - 1
    
    for asset_class in sorted_asset_classes:
        # Asset class name
        category_cell = sheet.cell(row=current_row, column=1, value=asset_class)
        category_cell.border = border
        category_cell.font = Font(bold=True)
        
        # Current Value - calculated from Global Portfolio (USD) data
        current_value = asset_class_values.get(asset_class, 0)
        current_value_cell = sheet.cell(row=current_row, column=4, value=current_value)
        current_value_cell.border = border
        current_value_cell.number_format = '$#,##0'
        
        # Current % = Current Value / Total Portfolio Value
        if total_portfolio_value > 0:
            current_pct = current_value / total_portfolio_value
        else:
            current_pct = 0
        current_pct_cell = sheet.cell(row=current_row, column=2, value=current_pct)
        current_pct_cell.border = border
        current_pct_cell.number_format = '0.0%'
        
        # Target %
        target_pct = target_allocations.get(asset_class, 0) / 100  # Default 0% for unknown classes
        target_cell = sheet.cell(row=current_row, column=3, value=target_pct)
        target_cell.border = border
        target_cell.number_format = '0.0%'
        
        # Target Value = Target % * Total Portfolio Value
        target_value = target_pct * total_portfolio_value
        target_value_cell = sheet.cell(row=current_row, column=5, value=target_value)
        target_value_cell.border = border
        target_value_cell.number_format = '$#,##0'
        
        # Difference = Current Value - Target Value
        difference = current_value - target_value
        diff_cell = sheet.cell(row=current_row, column=6, value=difference)
        diff_cell.border = border
        diff_cell.number_format = '$#,##0'
        
        # Action Required
        if abs(difference) > 1000:
            if difference > 0:
                action = f"Sell ${abs(difference):,.0f}"
            else:
                action = f"Buy ${abs(difference):,.0f}"
        else:
            action = "No action needed"
        
        action_cell = sheet.cell(row=current_row, column=7, value=action)
        action_cell.border = border
        
        current_row += 1
    
    # Position charts under the table data
    chart_start_row = data_end_row + 3
    
    # Create sorted helper data for pie chart (sort by current percentage descending)
    # Place helper data at column BA (53) onwards to avoid overlap with chart area
    helper_col = 53  # Column BA
    helper_start_row = 1  # Start at the top of the sheet, aligned with table headers
    
    # Collect asset class data with their current percentages for sorting
    asset_class_pct_data = []
    for row_idx in range(2, data_end_row + 1):
        asset_class_name = sheet.cell(row=row_idx, column=1).value
        current_pct_value = sheet.cell(row=row_idx, column=2).value or 0
        asset_class_pct_data.append((asset_class_name, current_pct_value, row_idx))
    
    # Sort by current percentage in descending order
    asset_class_pct_data.sort(key=lambda x: x[1], reverse=True)
    
    # Helper data headers
    sheet.cell(row=helper_start_row, column=helper_col, value="Sorted Asset Classes").font = Font(bold=True)
    sheet.cell(row=helper_start_row, column=helper_col + 1, value="Sorted Current %").font = Font(bold=True)
    
    # Write sorted helper data
    sorted_data_start = helper_start_row + 1
    for idx, (asset_class_name, current_pct, original_row) in enumerate(asset_class_pct_data):
        helper_row = sorted_data_start + idx
        sheet.cell(row=helper_row, column=helper_col, value=asset_class_name)
        # Reference the original cell to maintain formula linkage
        sheet.cell(row=helper_row, column=helper_col + 1, value=f"=B{original_row}")
    sorted_data_end = sorted_data_start + len(asset_class_pct_data) - 1
    
    # Add Target vs Current Allocation Bar Chart (positioned under the table)
    bar_chart = BarChart()
    bar_chart.title = "Target vs Current Allocation"
    bar_chart.style = 10
    bar_chart.y_axis.title = "Percentage"
    bar_chart.x_axis.title = "Asset Class"
    bar_chart.width = 15
    bar_chart.height = 10
    
    # Data for the bar chart (Current % and Target %)
    data = Reference(sheet, min_col=2, min_row=1, max_row=data_end_row, max_col=3)
    cats = Reference(sheet, min_col=1, min_row=2, max_row=data_end_row)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)
    
    # Add the bar chart under the table
    sheet.add_chart(bar_chart, f"A{chart_start_row}")
    
    # Add Current Allocation Pie Chart (positioned next to bar chart)
    pie = PieChart()
    pie.title = "Current Asset Allocation"
    pie.width = 12
    pie.height = 10
    
    # Use sorted helper data for pie chart (sorted by current % descending)
    if sorted_data_end >= sorted_data_start:
        labels = Reference(sheet, min_col=helper_col, min_row=sorted_data_start, max_row=sorted_data_end)
        pie_data = Reference(sheet, min_col=helper_col + 1, min_row=sorted_data_start, max_row=sorted_data_end)
        pie.add_data(pie_data)
        pie.set_categories(labels)
    
    # Add the pie chart next to the bar chart
    sheet.add_chart(pie, f"F{chart_start_row}")
    
    # Add conditional formatting for differences
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Apply conditional formatting to difference column
    sheet.conditional_formatting.add(f'F2:F{data_end_row}', 
                                   CellIsRule(operator='greaterThan', formula=['1000'], fill=red_fill))
    sheet.conditional_formatting.add(f'F2:F{data_end_row}', 
                                   CellIsRule(operator='lessThan', formula=['-1000'], fill=red_fill))
    sheet.conditional_formatting.add(f'F2:F{data_end_row}', 
                                   CellIsRule(operator='between', formula=['-1000', '1000'], fill=green_fill))
    
    # Auto-adjust column widths
    auto_adjust_column_widths(sheet)

def create_watchlist(sheet, header_font, header_fill, border, portfolio_data):
    """Create investment watchlist with real-time data from yfinance"""
    headers = ["Symbol", "Company Name", "Current Price", "52-Week High", "52-Week Low", 
               "P/E Ratio", "Dividend Yield %", "Market Cap (B)", "Target Price", "Notes", "Sector"]
    
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    
    # Use actual watchlist data from portfolio.json
    watchlist_stocks = []
    real_time_data = portfolio_data.get('real_time_data', {})
    
    for item in portfolio_data['watchlist']:
        symbol = item['symbol']
        currency = item.get('currency', '$')
        
        # Use real-time data if available
        if symbol in real_time_data:
            rt_data = real_time_data[symbol]
            company_name = rt_data['company_name']
            current_price = rt_data['current_price'] 
            high_52 = rt_data['52_week_high']
            low_52 = rt_data['52_week_low']
            pe_ratio = rt_data['pe_ratio'] or 0
            div_yield = rt_data['dividend_yield'] or 0
            market_cap_raw = rt_data['market_cap'] or 0
            market_cap = market_cap_raw / 1_000_000_000 if market_cap_raw else 0  # Convert to billions
            sector = rt_data['sector']
            
            # Use target price from watchlist item if provided
            target_price = item.get('target_price', current_price)
        else:
            # If no real-time data, use placeholder values
            company_name = item.get('name', symbol)
            current_price = 0
            high_52 = 0
            low_52 = 0
            pe_ratio = 0
            div_yield = 0
            market_cap = 0
            sector = 'Unknown'
            target_price = item.get('target_price', 0)
        
        watchlist_stocks.append((
            symbol,
            currency,
            company_name,
            current_price,
            high_52,
            low_52,
            pe_ratio,
            div_yield,
            market_cap,
            target_price,
            item.get('note', ''),
            sector
        ))
    
    for row, stock_data in enumerate(watchlist_stocks, 2):
        symbol, currency, company_name, current, high_52, low_52, pe, div_yield, market_cap, target_price, notes, sector = stock_data
        
        data = [symbol, company_name, current, high_52, low_52,
                pe, div_yield, market_cap, round(target_price, 2), notes, sector]
        
        for col, value in enumerate(data, 1):
            cell = sheet.cell(row=row, column=col, value=value)
            cell.border = border
            
            # Number formatting with currency support
            if col == 1:
                cell.font = Font(bold=True)
            elif col in [3, 4, 5, 9]:  # Price columns (Current Price, 52-Week High, 52-Week Low, Target Price)
                cell.number_format = get_currency_format(currency)
            elif col == 7:  # Dividend yield
                cell.number_format = '0.0%'
            elif col == 8:  # Market cap
                cell.number_format = '#,##0'
            elif col == 11:  # Sector column
                cell.font = Font(italic=True)
    
    # Auto-adjust column widths
    auto_adjust_column_widths(sheet)


if __name__ == "__main__":
    # Parse command-line arguments
    args = parse_arguments()
    
    # Get the appropriate portfolio JSON path
    portfolio_path = get_portfolio_json_path(use_test=args.test)
    
    # Print which file is being used
    print(f"{'='*60}")
    print(f"Using portfolio file: {portfolio_path.name}")
    print(f"{'='*60}")
    
    # Create the portfolio template
    create_investment_portfolio_template(portfolio_path)
